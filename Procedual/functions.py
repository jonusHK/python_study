# 2019.03.19 화요일

# get_raw_data(filename) -> dict
# dict = {'name1', score1, 'name2', score2, ...}

# get_average(...) -> integer
# get_variance(...) -> float<2>
# get_std_dev(...) -> float<2>
# get_evaluation(...) -> None
from openpyxl import load_workbook
from functools import reduce
from math import *

def get_raw_data(filename):
   raw_data = {}
   wb = load_workbook(filename)
   ws = wb.active
   g = ws.rows
   for name_cell, score_cell in g:
      raw_data[name_cell.value] = score_cell.value
   return raw_data
'''
# dictionary로 받지 않고 list로 받는 법 (get_scores 함수 불필요)
def get_raw_data(filename):
    stud_list = []
    wb = load_workbook(filename)
    ws = wb.active
    g = ws.rows
    for k, v in g:
        stud_list.append(v.value)
    return stud_list
'''

def get_average(scores):
   '''
   get_average(scores) -> float
   scores --> tuple lf scores
   return : float : 소수점 이하 두자리 반환
   '''
   avrg = reduce(lambda a, b : a + b, scores) / len(scores)
   return avrg

# 래핑함수 (wrapping function)
def get_scores(raw_data):
   '''
   get_scores(dict) -> tuple
   dict = {'name1' : score1, 'name2' : score2, ...}
   return : (score1, score2, score3)
   '''
   return tuple(raw_data.values())

def get_variance(scores, avrg):
   '''
   get_variance(scores, avrg) -> float
   scores --> tuple
   avrg --> 평균
   return : float : 소수점 이하 두 자리 반환
   '''
   variance = reduce(lambda result, a: result + (a - avrg) ** 2, scores, 0) / len(scores)
   return round(variance, 2)
   
def get_std_dev(variance):
   std_dev = sqrt(variance)
   return round(std_dev, 2)

def get_evaluation(avrg, variance, std_dev, total_avrg, sd = 20):
   print(f"평균: {avrg}, 분산: {variance}, 표준편차: {std_dev}")

   if avrg < total_avrg and variance > sd:
      print('성적이 너무 저조하고 학생들의 실력 차이가 너무 크다.')
   elif avrg > total_avrg and variance > sd:
      print('성적은 평균 이상이지만 학생들의 실력 차이가 크다. 주의 요망!')
   elif avrg < total_avrg and variance < sd:
      print('학생들의 실력 차이는 크지 않지만 성적이 너무 저조하다. 주의 요망!')
   elif avrg > total_avrg and variance < sd:
      print('성적도 평균 이상이고 학생들의 실력 차이도 크지 않다.')

# 실행하는 주체라면 실행하고, 아니라면 실행안하는 것
if __name__ == "__main__": 
   raw_data = get_raw_data('class_2-3.xlsx')
   #print(raw_data)
   scores = get_scores(raw_data)
   #print(scores)
   avrg = get_average(scores)
   #print(avrg)
   variance = get_variance(scores, avrg)
   #print(variance)
   std_dev = get_std_dev(variance)
   #print(std_dev)
   get_evaluation(avrg, variance, std_dev, 50)

